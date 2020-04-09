#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author    : jennifermq

import codecs
import json
import openpyxl

# 获取邮箱映射，以对姓名排序
wb = openpyxl.load_workbook('名单.xlsx')
names = wb.sheetnames
sheet = wb[names[0]]
name_list = []
name_index = 0
email_index = 8
email_map = {}
for index, row in enumerate(sheet.iter_rows()):
    try:
        if index == 0:
            index_list = [x.value for x in row]
            name_index = index_list.index("姓名")
            email_index = index_list.index("邮箱")
        else:
            name = row[name_index].value
            email = row[email_index].value
            email = email.split("@")[0]
            email_map[name.strip()] = email.strip()
    except:
        continue

# 将中奖名单写入文件
content = ""
with open('导出名单.txt', 'r', encoding='utf-8') as ifile:
    for line in ifile.readlines():
        try:
            name_list = json.loads(line)
            if len(name_list) == 0:
                content += "无\n"
            else:
                if isinstance(name_list[0], str):
                    name_list = sorted(name_list, key=lambda x: email_map.get(x, "a"))
                    content += ", ".join(name_list)
                    content += "\n\n"
                elif isinstance(name_list[0], list):
                    for index, group in enumerate(name_list):
                        g = sorted(group, key=lambda x: email_map.get(x, "a"))
                        content += "第 " + str(index + 1) + " 组\n"
                        content += ", ".join(g)
                        content += "\n\n"
                else:
                    content += line
        except:
            content += line

with codecs.open('中奖名单.txt', 'w', 'utf-8') as ofile:
    ofile.write(content)

