#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Author    : muqing
# Email     : mouqing@qianxin.com
# Department: anfu

import codecs
import json
import openpyxl

# 读取人员列表excel文件
wb = openpyxl.load_workbook('名单.xlsx')
names = wb.sheetnames
sheet = wb[names[0]]

name_list = []
index_list = []
name_index = 0
status_index = 3
for index, row in enumerate(sheet.iter_rows()):
    if index == 0:
        # 更新字段序号
        index_list = [x.value for x in row]
        name_index = index_list.index("姓名")
        status_index = index_list.index("是否在职")
    else:
        # 将在职人员添加到名字列表
        name = row[name_index].value
        status = row[status_index].value
        print(name, status)
        if name is None and status is None:
            continue
        if status.strip() == "在职中":
            name_list.append(name)

# 将名字列表写入js文件，抽奖程序可直接读取，不需要用ajax
with codecs.open('抽奖名单.js', 'w', 'utf-8') as ofile:
    content = "let namelist = "
    content += json.dumps(name_list, ensure_ascii=False)
    content += ";"
    ofile.write(content)
